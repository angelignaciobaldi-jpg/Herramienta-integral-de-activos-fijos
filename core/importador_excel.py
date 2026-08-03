"""Carga masiva de inventarios de activos fijos desde Excel.

Pensado para los levantamientos que el área ya tiene capturados en hojas de
cálculo (p. ej. el inventario de la clínica Oceánica), donde cada hoja es un área
y cada fila un activo (o VARIOS, ver abajo).

Estructura esperada de cada hoja (las columnas se detectan por su encabezado, sin
importar en qué fila empiece ni el orden):

    INSUMO · CANTIDAD · ETIQUETA · SERIE · RESPONSABLE · UBICACIÓN

Dos particularidades del formato real que este módulo resuelve:

1. **Una fila puede ser varios activos.** Cuando CANTIDAD > 1, la celda ETIQUETA
   trae todas las etiquetas juntas ("0048399/0048400/0048401", a veces separadas
   por saltos de línea). Cada etiqueta es un activo distinto en el SIPP, así que
   la fila se EXPANDE en un registro por etiqueta.
2. **La mayoría de los activos no tiene número de serie** (en el inventario real,
   ~82%). El identificador es la ETIQUETA (número de inventario); la serie queda
   como dato opcional (ver core/db.clave_levantamiento).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date as _date
from datetime import datetime as _datetime

import openpyxl

from . import db
from .empresas import ID_POR_EMPRESA
from .tipos_activo import ID_POR_NOMBRE, SITUACIONES, TIPOS_ACTIVO

# Encabezados que se buscan (en MAYÚSCULAS, sin acentos) -> campo interno.
# Las sábanas estandarizadas traen EMPRESA, SUCURSAL, ORIGEN (sitio) y AREA (área),
# que se usan por fila para autollenar el registro (ver _importar_hoja).
_ENCABEZADOS = {
    "INSUMO": "insumo",
    "CANTIDAD": "cantidad",
    "ETIQUETA": "etiqueta",
    "SERIE": "serie",
    "RESPONSABLE": "responsable",
    "UBICACION": "ubicacion",
    "EMPRESA": "empresa",
    "SUCURSAL": "sucursal",
    "ORIGEN": "origen",
    "AREA": "area",
}
# Campos del FORMULARIO DE ALTA en la plantilla completa (encabezado normalizado ->
# clave que consume el RPA en datos_json). Así el Excel es la base del alta: el
# usuario llena aquí lo que antes tecleaba en el formulario. El TIPO se traduce a
# id_tipo_activo aparte (no va en datos_json).
_ENCABEZADOS_ALTA = {
    "TIPO DE ACTIVO": "id_TipoActivo",
    "DESCRIPCION": "de_DescripcionActivo",
    "SITUACION": "id_Situacion",
    "COSTO": "im_Costo",
    "FACTURA": "nb_Factura",
    "PROVEEDOR": "nb_Proveedor",
    "EMPRESA COMPRA": "id_EmpresaAgregar",
    "SUCURSAL COMPRA": "id_SucursalAgregar",
    "GRUPO CENTRO DE COSTO": "id_GrupoCentroCosto",
    "CENTRO DE COSTO": "id_CentroCosto",
    "DEPARTAMENTO": "id_Departamento",
    "FECHA ADQUISICION": "FH_ADQUISICION",
    "FECHA GARANTIA": "FH_GARANTIA",
    "FECHA ASIGNACION": "FH_ASIGNACION",
    "MARCA": "marca",
    "MODELO": "modelo",
    "CLIENTE": "cliente",
    "PLACA": "placa",
}
# Detección combinada: encabezado -> destino (campo básico O clave del alta).
_HEADERS = {**_ENCABEZADOS, **_ENCABEZADOS_ALTA}

# Cuántos encabezados deben coincidir para dar una fila por "fila de encabezados".
_MIN_COINCIDENCIAS = 3
# Hasta qué fila se busca el encabezado (las hojas reales lo tienen entre la 6 y la 12).
_MAX_FILA_ENCABEZADO = 25
# Separadores con los que vienen varias etiquetas/series en una misma celda.
_RE_SEPARADORES = re.compile(r"[\/\n\r,;]+")


def _norm(texto) -> str:
    """Normaliza un encabezado: mayúsculas, sin acentos ni espacios sobrantes."""
    if texto is None:
        return ""
    t = str(texto).strip().upper()
    for a, b in (("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U"), ("Ñ", "N")):
        t = t.replace(a, b)
    return re.sub(r"\s+", " ", t)


def _partes(celda) -> list[str]:
    """Separa una celda que puede traer varios valores ('0048399/0048400')."""
    if celda is None:
        return []
    # Los números de etiqueta suelen venir como número: se evita el '.0' del float.
    if isinstance(celda, float) and celda.is_integer():
        celda = int(celda)
    return [p.strip() for p in _RE_SEPARADORES.split(str(celda)) if p.strip()]


@dataclass
class HojaDetectada:
    """Resultado del análisis de una hoja del archivo."""

    nombre: str
    fila_encabezado: int
    columnas: dict            # campo interno -> índice de columna (1-based)
    filas_datos: int          # filas con INSUMO
    activos_estimados: int    # tras expandir las etiquetas múltiples
    importable: bool = True
    motivo: str = ""


@dataclass
class ResultadoImportacion:
    """Estadísticas de una importación."""

    agregados: int = 0
    duplicados: int = 0
    sin_etiqueta: int = 0
    filas_leidas: int = 0
    errores: list = field(default_factory=list)


# Encabezados de la PLANTILLA de carga masiva (orden y nombres canónicos que el
# detector reconoce). Incluye TODOS los campos del formulario de alta para que el
# Excel sea la base del registro; lo no llenado se completa después en la herramienta.
PLANTILLA_ENCABEZADOS = [
    "INSUMO", "CANTIDAD", "ETIQUETA", "SERIE", "RESPONSABLE",
    "EMPRESA", "SUCURSAL", "UBICACION", "TIPO DE ACTIVO", "DESCRIPCION",
    "SITUACION", "COSTO", "FACTURA", "PROVEEDOR", "EMPRESA COMPRA",
    "SUCURSAL COMPRA", "GRUPO CENTRO DE COSTO", "CENTRO DE COSTO", "DEPARTAMENTO",
    "FECHA ADQUISICION", "FECHA GARANTIA", "FECHA ASIGNACION",
    "MARCA", "MODELO", "CLIENTE", "PLACA",
]


def generar_plantilla(ruta: str) -> str:
    """Crea en `ruta` un Excel plantilla de carga masiva: hoja «Activos» con TODOS
    los campos del alta y una hoja «Instrucciones». Devuelve la ruta escrita.

    Se deja SIN filas de datos para no importar ejemplos por error; el formato y el
    truco de varias etiquetas por fila se explican en la hoja de instrucciones.
    TIPO DE ACTIVO y SITUACION traen lista desplegable con los valores válidos."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activos"
    ws.append(PLANTILLA_ENCABEZADOS)
    for i, celda in enumerate(ws[1], 1):
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1F3A5F")
        celda.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = min(
            40, max(14, len(PLANTILLA_ENCABEZADOS[i - 1]) + 2))
    ws.freeze_panes = "A2"

    # Listas desplegables para TIPO DE ACTIVO y SITUACION (valores válidos del SIPP).
    def _validacion(valores, col_header):
        formula = '"' + ",".join(valores) + '"'
        if len(formula) > 255:  # límite de Excel para listas embebidas
            return
        col = get_column_letter(PLANTILLA_ENCABEZADOS.index(col_header) + 1)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}1000")

    _validacion(list(TIPOS_ACTIVO.values()), "TIPO DE ACTIVO")
    _validacion(list(SITUACIONES.values()), "SITUACION")

    ins = wb.create_sheet("Instrucciones")
    ins.column_dimensions["A"].width = 100
    guia = [
        "CARGA MASIVA DE ACTIVOS — INSTRUCCIONES",
        "",
        "Captura un activo por fila en la hoja «Activos». El Excel es la base del "
        "registro: lo que llenes aquí es lo que usará el alta automática (RPA); lo "
        "que dejes vacío se completa después en la herramienta.",
        "",
        "IDENTIFICACIÓN",
        "• INSUMO (obligatorio): nombre del insumo tal como está en el SIPP (se "
        "resuelve a su clave para el RPA; si no se encuentra, se elige en la ficha).",
        "• CANTIDAD: cuántas piezas iguales. Si es más de 1, pon todas sus ETIQUETAS "
        "en la misma celda separadas por «/» (ej. 0048399/0048400/0048401): cada "
        "etiqueta se registra como un activo.",
        "• ETIQUETA: número(s) de inventario (identificador principal).",
        "• SERIE: número de serie (opcional). Si hay varias, sepáralas por «/» en el "
        "mismo orden que las etiquetas.",
        "• TIPO DE ACTIVO: elige de la lista desplegable.",
        "• DESCRIPCION, SITUACION (lista desplegable).",
        "",
        "COMPRA",
        "• COSTO, FACTURA (folio), PROVEEDOR, EMPRESA COMPRA, SUCURSAL COMPRA.",
        "• GRUPO CENTRO DE COSTO, CENTRO DE COSTO, DEPARTAMENTO: tal como aparecen en "
        "el catálogo del SIPP de la empresa.",
        "• FECHAS (ADQUISICION / GARANTIA / ASIGNACION): formato DD/MM/AAAA.",
        "",
        "RESGUARDO",
        "• RESPONSABLE: empleado resguardante (se resuelve a su id para el RPA).",
        "• EMPRESA y SUCURSAL: la empresa/sucursal del activo. Si las dejas vacías, se "
        "usan las del selector de la pantalla al importar.",
        "• UBICACION: ubicación física.",
        "",
        "CARACTERÍSTICAS (según el tipo): MARCA, MODELO, CLIENTE, PLACA.",
        "",
        "No cambies los nombres de los encabezados de la hoja «Activos».",
    ]
    for i, linea in enumerate(guia, 1):
        ins.cell(row=i, column=1, value=linea)
    ins["A1"].font = Font(bold=True, size=13)

    wb.save(ruta)
    return ruta


def analizar(ruta: str) -> list[HojaDetectada]:
    """Analiza el archivo y devuelve qué hojas se pueden importar y cuántos
    activos saldrían de cada una (ya expandidos). No modifica nada."""
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    try:
        return [_analizar_hoja(wb[nombre]) for nombre in wb.sheetnames]
    finally:
        wb.close()


def _analizar_hoja(ws) -> HojaDetectada:
    fila_hdr, columnas = _detectar_encabezado(ws)
    if not columnas or "insumo" not in columnas:
        return HojaDetectada(
            ws.title, 0, {}, 0, 0, importable=False,
            motivo="No se encontraron los encabezados esperados (INSUMO, ETIQUETA…).")
    filas = activos = 0
    for valores in _filas_datos(ws, fila_hdr, columnas):
        filas += 1
        activos += max(1, len(_partes(valores.get("etiqueta"))))
    return HojaDetectada(ws.title, fila_hdr, columnas, filas, activos)


def _match_header(texto: str) -> "str | None":
    """Destino (campo básico o clave del alta) de un encabezado. Prefiere el match
    EXACTO; si no, el encabezado más específico contenido en el texto (el más largo),
    para que 'EMPRESA COMPRA' no lo capture 'EMPRESA' ni 'GRUPO CENTRO DE COSTO' a
    'CENTRO DE COSTO'."""
    if texto in _HEADERS:
        return _HEADERS[texto]
    candidatos = [(h, d) for h, d in _HEADERS.items() if h in texto]
    if candidatos:
        return max(candidatos, key=lambda hd: len(hd[0]))[1]
    return None


def _detectar_encabezado(ws) -> tuple[int, dict]:
    """Busca la fila de encabezados y mapea destino (campo básico o clave del alta)
    -> índice de columna."""
    for i, fila in enumerate(ws.iter_rows(min_row=1, max_row=_MAX_FILA_ENCABEZADO,
                                          values_only=True), 1):
        columnas = {}
        for j, celda in enumerate(fila, 1):
            texto = _norm(celda)
            if not texto:
                continue
            destino = _match_header(texto)
            if destino and destino not in columnas:
                columnas[destino] = j
        if len(columnas) >= _MIN_COINCIDENCIAS and "insumo" in columnas:
            return i, columnas
    return 0, {}


def _fmt_valor(clave: str, valor) -> str:
    """Normaliza el valor de una celda del alta a texto (fechas -> DD/MM/AAAA)."""
    if valor is None:
        return ""
    if clave.startswith("FH_") and isinstance(valor, (_datetime, _date)):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return str(valor).strip()


def _resolver_insumo(nombre: str, id_empresa, cache: dict) -> "int | None":
    """Clave (id) del insumo por nombre exacto en la caché del SIPP; None si no está.
    Busca primero en la empresa y, si no, en el catálogo global."""
    n = _norm(nombre)
    if not n:
        return None
    if (n, id_empresa) in cache:
        return cache[(n, id_empresa)]
    encontrado = None
    candidatos = (db.buscar_insumos(nombre, empresa_id=id_empresa, limite=25)
                  or db.buscar_insumos(nombre, limite=25))
    for ins in candidatos:
        if _norm(ins.nombre) == n:
            encontrado = ins.id_insumo
            break
    cache[(n, id_empresa)] = encontrado
    return encontrado


def _resolver_empleado(nombre: str, cache: dict) -> "int | None":
    """id del empleado por nombre exacto en la caché del SIPP; None si no está."""
    n = _norm(nombre)
    if not n:
        return None
    if n in cache:
        return cache[n]
    encontrado = None
    for emp in db.buscar_empleados(nombre, limite=25):
        if _norm(emp.nombre) == n:
            encontrado = emp.id_empleado
            break
    cache[n] = encontrado
    return encontrado


def _filas_datos(ws, fila_hdr: int, columnas: dict):
    """Itera las filas con dato, ya mapeadas a {campo: valor}."""
    for fila in ws.iter_rows(min_row=fila_hdr + 1, max_row=ws.max_row, values_only=True):
        insumo = fila[columnas["insumo"] - 1] if columnas["insumo"] - 1 < len(fila) else None
        if insumo is None or not str(insumo).strip():
            continue
        yield {campo: (fila[idx - 1] if idx - 1 < len(fila) else None)
               for campo, idx in columnas.items()}


def importar(ruta: str, hojas: list[str], empresa: str = "", sucursal: str = "",
             departamento: str = "", progreso=None) -> ResultadoImportacion:
    """Importa las `hojas` indicadas del archivo al levantamiento.

    Cada fila se expande en un registro por ETIQUETA. EMPRESA, SUCURSAL y
    DEPARTAMENTO (columna AREA) se autollenan por fila desde el archivo cuando
    existen esas columnas; si la hoja no las trae, se usan los argumentos
    `empresa`/`sucursal`/`departamento` como respaldo. La ubicación combina el
    sitio (ORIGEN) con la UBICACIÓN. En sábanas antiguas sin columna SUCURSAL,
    ORIGEN se usa como sucursal (compatibilidad). El TIPO de activo se deja vacío
    para asignarlo después desde la herramienta.

    `progreso(hecho, total, hoja)`: callback opcional para reflejar el avance.
    """
    res = ResultadoImportacion()
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    try:
        total = len(hojas)
        for n, nombre in enumerate(hojas, 1):
            if nombre not in wb.sheetnames:
                res.errores.append(f"La hoja '{nombre}' no existe en el archivo.")
                continue
            ws = wb[nombre]
            fila_hdr, columnas = _detectar_encabezado(ws)
            if not columnas or "insumo" not in columnas:
                res.errores.append(f"'{nombre}': no se detectaron los encabezados.")
                continue
            if progreso:
                progreso(n, total, nombre)
            _importar_hoja(ws, fila_hdr, columnas, empresa, sucursal,
                           departamento, res)
    finally:
        wb.close()
    return res


def _importar_hoja(ws, fila_hdr: int, columnas: dict, empresa: str, sucursal: str,
                   departamento: str, res: ResultadoImportacion) -> None:
    """Arma los registros de la hoja (expandiendo etiquetas) y los inserta EN
    LOTE: con miles de filas, una transacción por registro es muchísimo más lenta."""
    registros = []
    cache_insumo: dict = {}
    cache_empleado: dict = {}
    for valores in _filas_datos(ws, fila_hdr, columnas):
        res.filas_leidas += 1
        insumo = str(valores.get("insumo") or "").strip()
        etiquetas = _partes(valores.get("etiqueta"))
        series = _partes(valores.get("serie"))
        responsable = str(valores.get("responsable") or "").strip()
        # Empresa/sucursal se autollenan por fila desde sus columnas; si la hoja no
        # trae la columna (o la celda está vacía) se usa el valor de la UI.
        emp_fila = str(valores.get("empresa") or "").strip() or empresa
        # DEPARTAMENTO: ahora hay columna dedicada (clave id_Departamento); si no
        # viene, respaldo al del selector de la UI.
        dep_fila = _fmt_valor("id_Departamento", valores.get("id_Departamento")) or departamento
        origen = str(valores.get("origen") or "").strip()
        ubic = str(valores.get("ubicacion") or "").strip()
        suc_col = str(valores.get("sucursal") or "").strip()
        if suc_col:
            # Formato con columna SUCURSAL: esa es la sucursal; ORIGEN (el sitio)
            # enriquece la ubicación para no perderlo.
            suc_fila = suc_col
            ubic_fila = " — ".join(p for p in (origen, ubic) if p)
        else:
            # Formato anterior sin columna SUCURSAL: ORIGEN es la sucursal.
            suc_fila = origen or sucursal
            ubic_fila = ubic

        # --- Campos del ALTA (plantilla completa) -> datos_json ---------------
        id_empresa = ID_POR_EMPRESA.get(emp_fila)
        tipo_nombre = _fmt_valor("id_TipoActivo", valores.get("id_TipoActivo"))
        id_tipo = ID_POR_NOMBRE.get(tipo_nombre) if tipo_nombre else None
        datos: dict = {}
        for clave in _ENCABEZADOS_ALTA.values():
            if clave == "id_TipoActivo":   # va como id_tipo_activo, no en datos
                continue
            v = _fmt_valor(clave, valores.get(clave))
            if v:
                datos[clave] = v
        # Lo que el formulario también guarda en datos y aquí ya conocemos.
        if insumo:
            datos["nb_NombreInsumo"] = insumo
        if responsable:
            datos["nb_Empleado"] = responsable
        if ubic_fila:
            datos.setdefault("de_Ubicacion", ubic_fila)
        if dep_fila:
            datos.setdefault("id_Departamento", dep_fila)
        # Ids del SIPP para que el RPA seleccione insumo/empleado por id (si no se
        # resuelven, se dejan para elegirlos en la ficha).
        id_ins = _resolver_insumo(insumo, id_empresa, cache_insumo)
        if id_ins:
            datos["id_InsumoOrigen"] = str(id_ins)
        id_emp = _resolver_empleado(responsable, cache_empleado)
        if id_emp:
            datos["id_EmpleadoResguardo"] = str(id_emp)

        if not etiquetas:
            # Sin etiqueta: se guarda un único registro (se identificará por
            # insumo + serie) y se reporta para que el área lo revise.
            res.sin_etiqueta += 1
            etiquetas = [""]

        for i, etiqueta in enumerate(etiquetas):
            # La serie se aparea posicionalmente con la etiqueta; si hay menos
            # series que etiquetas, las restantes quedan sin serie (lo normal:
            # una fila de 10 sillas trae 10 etiquetas y ninguna serie).
            serie_i = series[i] if i < len(series) else ""
            datos_fila = dict(datos)
            if serie_i:
                datos_fila["nu_Serie"] = serie_i
            registros.append({
                "nombre_insumo": insumo,
                "etiqueta": etiqueta,
                "no_serie": serie_i,
                "responsable": responsable,
                "ubicacion": ubic_fila,
                "empresa": emp_fila,
                "sucursal": suc_fila,
                "departamento": dep_fila,
                "id_tipo_activo": id_tipo,
                "datos": datos_fila or None,
            })

    agregados, duplicados = db.guardar_levantamiento_lote(registros)
    res.agregados += agregados
    res.duplicados += duplicados
