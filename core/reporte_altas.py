"""Reporte de un proceso de altas en el SIPP (RPA).

Al terminar el lote de altas, la herramienta arma un reporte con las estadísticas
generales (altas realizadas, pendientes) y las observaciones por activo (p. ej.
"no se encontró el insumo"), y permite exportarlo a Excel.
"""

from __future__ import annotations

from datetime import datetime

import openpyxl

# Estatus posibles de cada activo en el reporte.
ALTA = "Alta realizada"
PENDIENTE = "Pendiente"

_ENCABEZADOS = ["Insumo", "Etiqueta", "No. de serie", "Estatus", "Observación"]


def resumen_altas(filas: list[dict], detenido: bool = False) -> dict:
    """Estadísticas generales a partir de las filas del reporte."""
    realizadas = sum(1 for f in filas if f.get("estatus") == ALTA)
    total = len(filas)
    return {"total": total, "realizadas": realizadas,
            "pendientes": total - realizadas, "detenido": detenido}


def generar_reporte_altas(ruta: str, filas: list[dict],
                          detenido: bool = False,
                          errores_generales: "list[str] | None" = None) -> str:
    """Escribe en `ruta` un Excel con el resumen y el detalle por activo.
    Devuelve la ruta escrita. `filas`: dicts con insumo, etiqueta, serie, estatus,
    observacion."""
    from openpyxl.styles import Alignment, Font, PatternFill

    res = resumen_altas(filas, detenido)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de altas"

    titulo = ws.cell(row=1, column=1, value="Reporte de altas en el SIPP")
    titulo.font = Font(bold=True, size=14)
    ws.cell(row=2, column=1,
            value="Generado: " + datetime.now().strftime("%d/%m/%Y %H:%M"))
    resumen_lineas = [
        ("Total de activos procesados", res["total"]),
        ("Altas realizadas", res["realizadas"]),
        ("Pendientes", res["pendientes"]),
    ]
    if detenido:
        resumen_lineas.append(("Proceso", "DETENIDO por el usuario"))
    fila = 4
    for etiqueta, valor in resumen_lineas:
        ws.cell(row=fila, column=1, value=etiqueta).font = Font(bold=True)
        ws.cell(row=fila, column=2, value=valor)
        fila += 1

    if errores_generales:
        fila += 1
        ws.cell(row=fila, column=1, value="Observaciones generales").font = Font(bold=True)
        fila += 1
        for err in errores_generales:
            ws.cell(row=fila, column=1, value="• " + err)
            fila += 1

    # Detalle por activo.
    fila += 1
    for i, encabezado in enumerate(_ENCABEZADOS, 1):
        celda = ws.cell(row=fila, column=i, value=encabezado)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1F3A5F")
        celda.alignment = Alignment(horizontal="center")
    fila += 1
    for f in filas:
        ws.cell(row=fila, column=1, value=f.get("insumo", ""))
        ws.cell(row=fila, column=2, value=f.get("etiqueta", ""))
        ws.cell(row=fila, column=3, value=f.get("serie", ""))
        ws.cell(row=fila, column=4, value=f.get("estatus", ""))
        ws.cell(row=fila, column=5, value=f.get("observacion", ""))
        fila += 1

    anchos = [40, 16, 22, 16, 60]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    wb.save(ruta)
    return ruta
